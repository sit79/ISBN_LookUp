from openpyxl import load_workbook
import webbrowser

workbook = load_workbook(filename="isbn.xlsx")
sheet = workbook["Books"]

for i in range(3, 5):
    isbn = sheet.cell(row=i, column=2).value
    url = 'https://isbnsearch.org/isbn/{isbn}'
    webbrowser.open(url, new=0, autoraise=True)
